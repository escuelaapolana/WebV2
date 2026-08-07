from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

NAVY="2E4256"; AMAR="FDF3E3"; LINEA="D4CBB9"; CREMA="F1EADC"; AZULS="EAF2F9"
F="Arial"
fina = Side("thin", color=LINEA)
BORDE = Border(left=fina, right=fina, top=fina, bottom=fina)

GRUPOS = [
    ("Máster", ["Calle 1 · Iniciación", "Calle 2 · Desarrollo", "Calle 3 · Perfeccionamiento"]),
    ("Escuela 6-9", ["Calle 1 · Iniciación", "Calle 2 · Desarrollo"]),
    ("Escuela 10-15", ["Calle 1 · Iniciación", "Calle 2 · Desarrollo", "Calle 3 · Perfeccionamiento"]),
]
FILAS = ["Calentamiento", "Parte principal", "Vuelta a la calma", "Observaciones"]
DIAS = ["Lunes","Martes","Miércoles","Jueves","Viernes","Sábado","Domingo"]

# El ejemplo va POR GRUPO: en el máster la calle de iniciación son adultos que
# nadan menos, no niños con churro. Poner el juego de los peques en el máster
# enseña justo lo que no hay que escribir ahí.
EJEMPLO = {
 "Máster": {
   "Calentamiento": ["200 crol suave\n4x50 c/TABLA", "300 crol suave\n4x50 estilos c/ALETAS", "400 variado\n4x50 pies c/TABLA"],
   "Parte principal": ["6x50 A2 rec 20\"\n4x50 ritmo medio rec 20\"", "6x100 A2 rec 20\"\n8x50 progresivos rec 20\"", "8x100 A2 rec 20\"\n4x50 fuerte rec 30\""],
   "Vuelta a la calma": ["150 suaves", "200 suaves estilos", "300 suaves elección"],
   "Observaciones": ["Sin prisa, técnica por encima del ritmo", "Mantener la técnica al cansarse", "Última serie a ritmo de competición"],
 },
 "Escuela 6-9": {
   "Calentamiento": ["5 m salto + flecha\n5 m patada dorsal con churro\n15 m deslizamiento haciendo burbujas", "25 punto muerto\n25 rozando uñas"],
   "Parte principal": ["25 m salto + flecha + recoger pelota\n25 m patada crol con tabla\n25 m pasar bajo el aro", "6x25 crol RC3\n4x25 espalda"],
   "Vuelta a la calma": ["Juego: rescate del tesoro", "Juego: los tiburones"],
   "Observaciones": ["Mucho material de flotación", "Sin material"],
 },
 "Escuela 10-15": {
   "Calentamiento": ["200 (100) suaves c/ALETAS", "400 suaves c/ALETAS\n4x50 rec 15\"", "400 suaves c/ALETAS\n4x50 rec 15\""],
   "Parte principal": ["Escalera: 12,5 · 25 · 50 · 25 · 12,5", "Escalera: 25 · 50 · 100 · 50 · 25\n2 vueltas, la 2.ª c/ALETAS", "Escalera: 25 · 50 · 100 · 50 · 25\n2 vueltas, la 2.ª c/ALETAS"],
   "Vuelta a la calma": ["8x25 técnica rec 20\"", "8x50 técnica rec 20\"", "8x50 técnica rec 20\""],
   "Observaciones": ["Mantener la técnica aunque aumente el cansancio", "", ""],
 },
}

wb = Workbook(); wb.remove(wb.active)

for nombre, calles in GRUPOS:
    ws = wb.create_sheet(nombre)
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 20
    for i in range(len(calles)):
        ws.column_dimensions[chr(66+i)].width = 42

    ws["A1"] = "PLANIFICACIÓN SEMANAL · " + nombre.upper()
    ws["A1"].font = Font(name=F, size=15, bold=True, color=NAVY)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=1+len(calles))

    ws["A2"] = "Semana del lunes:"; ws["A2"].font = Font(name=F, size=10, bold=True, color=NAVY)
    ws["B2"] = "2026-08-10"
    ws["B2"].fill = PatternFill("solid", fgColor=AMAR); ws["B2"].border = BORDE
    ws["B2"].font = Font(name=F, size=10)
    if len(calles) > 1:
        ws.cell(row=2, column=3, value="Piscina:").font = Font(name=F, size=10, bold=True, color=NAVY)
        c = ws.cell(row=2, column=4 if len(calles) > 2 else 3)
    ws["A3"] = ("Una columna por calle, y cada calle es un nivel. Rellena solo lo crema. "
                "Si un día no se entrena, déjalo en blanco.")
    ws["A3"].font = Font(name=F, size=9, italic=True, color="6E6656")
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=1+len(calles))

    r = 5
    for d, dia in enumerate(DIAS):
        # cabecera del día
        cd = ws.cell(row=r, column=1, value=dia)
        cd.font = Font(name=F, size=12, bold=True, color="FFFFFF")
        cd.fill = PatternFill("solid", fgColor=NAVY)
        cd.alignment = Alignment(vertical="center")
        for j in range(len(calles)):
            cc = ws.cell(row=r, column=2+j, value=calles[j])
            cc.font = Font(name=F, size=10, bold=True, color="FFFFFF")
            cc.fill = PatternFill("solid", fgColor=NAVY)
            cc.alignment = Alignment(vertical="center", horizontal="center")
        ws.row_dimensions[r].height = 22
        r += 1
        # piscina y hora del día
        ws.cell(row=r, column=1, value="Piscina · hora").font = Font(name=F, size=9, color="6E6656")
        pc = ws.cell(row=r, column=2)
        pc.fill = PatternFill("solid", fgColor=AMAR); pc.border = BORDE
        pc.font = Font(name=F, size=10)
        if d == 0: pc.value = "25 m · 19:00"
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=1+len(calles))
        r += 1
        # las cuatro filas
        for et in FILAS:
            ce = ws.cell(row=r, column=1, value=et)
            ce.font = Font(name=F, size=10, bold=(et != "Observaciones"), color=NAVY)
            ce.fill = PatternFill("solid", fgColor=CREMA if et != "Observaciones" else AZULS)
            ce.alignment = Alignment(vertical="top"); ce.border = BORDE
            for j in range(len(calles)):
                cel = ws.cell(row=r, column=2+j)
                cel.fill = PatternFill("solid", fgColor=AMAR)
                cel.border = BORDE
                cel.font = Font(name=F, size=10)
                cel.alignment = Alignment(vertical="top", wrap_text=True)
                ej = EJEMPLO.get(nombre, {}).get(et, [])
                if d == 0 and j < len(ej) and ej[j]:
                    cel.value = ej[j]
                    cel.font = Font(name=F, size=10, italic=True, color="6E6656")
            ws.row_dimensions[r].height = 74 if et != "Observaciones" else 40
            r += 1
        r += 1   # un renglón de aire entre días

    if len(calles) > 1:
        dv = DataValidation(type="list", formula1='"25 m,50 m,50 m exterior"', allow_blank=True)
        ws.add_data_validation(dv)

    ws.freeze_panes = "B5"

# ---- hoja de instrucciones ----
g = wb.create_sheet("Cómo se rellena")
g.sheet_view.showGridLines = False
g.column_dimensions["A"].width = 26; g.column_dimensions["B"].width = 88
g["A1"] = "CÓMO SE RELLENA"; g["A1"].font = Font(name=F, size=14, bold=True, color=NAVY)
BL = [
 ("UNA HOJA POR GRUPO",""),
 ("Abajo, las pestañas","Máster · Escuela 6-9 · Escuela 10-15. Cada una es su semana."),
 ("",""),
 ("UNA COLUMNA POR CALLE",""),
 ("Y cada calle es un nivel","Calle 1 iniciación · Calle 2 desarrollo · Calle 3 perfeccionamiento."),
 ("Si todas hacen lo mismo","Escríbelo en las tres, o solo en la 2 y pon «igual» en las otras. Como te sea cómodo."),
 ("⚠️ La iniciación no es la mitad","Escribe sus metros reales. Si no los sabes, déjalo vacío: vacío es una respuesta, un número inventado no."),
 ("",""),
 ("CUATRO FILAS POR DÍA",""),
 ("Calentamiento","Parte principal · Vuelta a la calma · Observaciones."),
 ("Piscina · hora","Encima de las cuatro, para el día entero."),
 ("Un día sin entrenar","Se deja en blanco y ya."),
 ("",""),
 ("LOS CÓDIGOS DE SIEMPRE",""),
 ("Se escriben igual","A1 A2 A3 intensidad · RC2 RC3 RC4 respiración · C/ALETAS con material · rec 20\" la salida."),
 ("Estilos","crol · esp · br · mar · estilos"),
]
r = 3
for a,b in BL:
    if a and not b:
        c = g.cell(row=r, column=1, value=a)
        c.font = Font(name=F, size=11, bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=NAVY)
        g.cell(row=r, column=2).fill = PatternFill("solid", fgColor=NAVY)
    elif a:
        g.cell(row=r, column=1, value=a).font = Font(name=F, size=10, bold=True, color=NAVY)
        cb = g.cell(row=r, column=2, value=b)
        cb.font = Font(name=F, size=10); cb.alignment = Alignment(wrap_text=True, vertical="top")
    r += 1

wb.save("/tmp/plantilla-semana-natacion-calles.xlsx")
print("hojas:", wb.sheetnames)
