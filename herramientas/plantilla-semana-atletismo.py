#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Genera assets/descargas/plantilla-semana-atletismo.xlsx

El Excel que se descarga el entrenador de atletismo desde su portal, rellena en
Google Sheets y sube como CSV en «Planificar semana → Importar un archivo».

Por qué un Excel y no un formato de texto: un Excel se explica solo. El que
venga el año que viene lo abre, ve las columnas y sabe qué tiene que hacer. Un
formato de texto hay que enseñárselo a cada uno, y se olvida.

Las CABECERAS NO SE TOCAN a la ligera: el importador del portal las reconoce por
su nombre (ALIAS_CSV y claveDeCabecera, en portal/entrenador/index.html). Si se
cambia una aquí sin cambiarla allí, esa columna entera se cae al subir el archivo
y no avisa nadie. Ya pasó con la plantilla de natación.

Se ejecuta a mano cuando haya que cambiar algo:
    python3 herramientas/plantilla-semana-atletismo.py
"""

import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

# Los mismos colores que la plantilla de natación: las dos son de la misma casa.
AZUL   = '2E4256'   # cabeceras
CREMA  = 'FDF3E3'   # lo que se rellena
CIELO  = 'EAF2F9'   # tablas calculadas
GRIS   = '6E6656'   # el ejemplo, en cursiva
AMBAR  = 'B96F09'   # avisos
AMBAR2 = '8A5307'

FILA_1 = 6      # primera fila de datos
FILA_N = 205    # última

# (cabecera, ancho). La cabecera es contrato con el portal: ver aviso de arriba.
COLUMNAS = [
    ('Día',                                            12),
    ('Título de la sesión',                            26),
    ('Nota del día (la leen el atleta y su familia)',  38),
    ('Hora',                                            7),
    ('Lugar',                                          22),
    ('Tipo de sesión',                                 14),
    ('Papel del día',                                  16),
    ('Bloque',                                         22),
    ('Para quién (atleta o subgrupo)',                 24),
    ('Ejercicio',                                      40),
    ('Series',                                          8),
    ('Distancia (m)',                                  13),
    ('Ritmo',                                          14),
    ('Rec',                                            10),
    ('Calzado',                                        10),
    ('Carga (% RM)',                                   13),
    ('Observaciones',                                  36),
    ('Detalle',                                        16),
]
COL = {c[0]: get_column_letter(i + 1) for i, c in enumerate(COLUMNAS)}

DIAS = ['Lunes', 'Martes', 'Miércoles', 'Jueves', 'Viernes', 'Sábado', 'Domingo']

# Las listas desplegables. Todos los valores los entiende el importador tal cual
# (SINON_TIPO, SINON_ROL y SINON_CALZADO en el portal).
TIPOS   = ['Pista', 'Gimnasio', 'Rodaje', 'Activación', 'Competición', 'Descanso']
PAPELES = ['Calidad fuerte', 'Secundaria', 'Activación', 'Último toque 48h', 'Descarga', 'Competición']
BLOQUES = ['Parte inicial', 'Movilidad', 'Técnica', 'Velocidad máxima', 'Parte principal',
           'Salto', 'Vallas', 'Fuerza / Potencia', 'Tren superior', 'Potenciación neural',
           'Transferencia', 'Core', 'Vuelta a la calma']
CALZADOS = ['Z', 'C', 'ZG', 'Descalzo']

# La semana de ejemplo. Sale de las planificaciones de verdad del club (los PDF
# de la S8 y las hojas de Juan/Ander), y enseña a propósito las DOS formas de
# trabajar: el lunes todos juntos pero con parte principal distinta por atleta, y
# el miércoles y el viernes el grupo entero haciendo lo mismo.
EJEMPLO = [
    # día, título, nota, hora, lugar, tipo, papel, bloque, para quién,
    # ejercicio, series, distancia, ritmo, rec, calzado, carga, obs, detalle
    ('Lunes', 'Velocidad máxima', 'Frescos tras el fin de semana. Volumen mínimo y calidad absoluta.',
     '19:00', 'Estadio Joaquín Villar', 'Pista', 'Calidad fuerte',
     'Parte inicial', '', 'Wickets bajos (frecuencia y postura)', 3, 20, '', "1'", 'Z', '', '3-5 apoyos desde salida', ''),
    ('Lunes', '', '', '', '', '', '', 'Parte inicial', '', 'Progresivos', 4, 50, '70-85%', "2'", 'Z', '', '', ''),
    ('Lunes', '', '', '', '', '', '', 'Parte principal', 'Juan', 'Salidas desde tacos', 4, 20, '95%', "3'", 'C', '',
     'Primera zancada dirigida. Tronco estable', ''),
    ('Lunes', '', '', '', '', '', '', 'Parte principal', 'Juan', 'Aceleración', 2, 40, '95%', "4'", 'C', '', '', ''),
    ('Lunes', '', '', '', '', '', '', 'Parte principal', 'Ander', 'Salidas desde tacos', 4, 20, '95%', "3'", 'C', '',
     'Mantener empuje, no retroceder fémur', ''),
    ('Lunes', '', '', '', '', '', '', 'Parte principal', 'Ander', 'Curva', 2, 120, '90%', "6'", 'C', '',
     'Buena transición. 14.5 s', ''),
    ('Lunes', '', '', '', '', '', '', 'Vuelta a la calma', '', 'Trote suave + estiramientos', '', '', '', '', 'Z', '', '', "8'"),

    ('Miércoles', 'Series de 300', 'Ritmos cómodos. Mantener la mecánica hasta el final.',
     '19:00', 'Estadio Joaquín Villar', 'Pista', 'Secundaria',
     'Parte inicial', '', 'Progresivos', 3, 60, '', "1'", 'Z', '', '', ''),
    ('Miércoles', '', '', '', '', '', '', 'Parte principal', '', 'Series de 300', 2, 300, '90-92%', "5'", 'C', '',
     'Ritmo controlado', ''),
    ('Miércoles', '', '', '', '', '', '', 'Parte principal', '', 'Series de 200', 2, 200, '92-94%', "4'", 'C', '', '', ''),
    ('Miércoles', '', '', '', '', '', '', 'Vuelta a la calma', '', 'Trote suave', '', '', '', '', 'Z', '', '', "8'"),

    ('Viernes', 'Fuerza y potencia', 'Sin llegar a fallo. Todo movido rápido.',
     '18:30', 'Gimnasio del club', 'Gimnasio', 'Secundaria',
     'Fuerza / Potencia', '', 'Cargada completa', 3, 2, '', "3'", 'ZG', '80% RM', 'Movida rapidísimo', ''),
    ('Viernes', '', '', '', '', '', '', 'Fuerza / Potencia', '', 'Sentadilla', 4, 6, '', "3'", 'ZG', '75% RM',
     'Bajada controlada', ''),
    ('Viernes', '', '', '', '', '', '', 'Core', '', 'Plancha', 3, '', '', "45\"", 'Descalzo', '', '', ''),

    ('Sábado', 'Descanso', '', '', '', 'Descanso', 'Descarga', '', '', '', '', '', '', '', '', '', '', ''),
]


def escribir(ws, celda, valor, fuente=None, relleno=None, alineacion=None, borde=None):
    c = ws[celda]
    c.value = valor
    if fuente:     c.font = fuente
    if relleno:    c.fill = relleno
    if alineacion: c.alignment = alineacion
    if borde:      c.border = borde
    return c


def hoja_semana(wb):
    ws = wb.create_sheet('Semana')
    ws.sheet_view.showGridLines = False

    f_titulo  = Font(name='Arial', bold=True, size=16, color=AZUL)
    f_etiq    = Font(name='Arial', bold=True, size=10, color=AZUL)
    f_pista   = Font(name='Arial', italic=True, size=9, color=GRIS)
    f_cab     = Font(name='Arial', bold=True, size=10, color='FFFFFF')
    f_dato    = Font(name='Arial', size=10)
    f_ejemplo = Font(name='Arial', italic=True, size=10, color=GRIS)
    f_marca   = Font(name='Arial', bold=True, size=9, color=AMBAR)

    r_crema = PatternFill('solid', fgColor=CREMA)
    r_azul  = PatternFill('solid', fgColor=AZUL)
    borde   = Border(*[Side(style='thin', color='D8CDB8')] * 4)
    arriba  = Alignment(vertical='top', wrap_text=True)

    escribir(ws, 'A1', 'PLANIFICACIÓN SEMANAL · ATLETISMO', f_titulo)
    ws.row_dimensions[1].height = 26

    # La cabecera de la semana, tal como la escriben las planificaciones del club:
    # grupo, semana, bloque de la temporada y competición de referencia.
    for celda, texto in [('A2', 'Grupo:'), ('C2', 'Semana del lunes:'), ('F2', 'Bloque de la temporada:'),
                         ('H2', 'Competición de referencia:')]:
        escribir(ws, celda, texto, f_etiq)
    for celda, texto in [('B2', 'Velocistas'), ('D2', '2026-08-10'), ('G2', 'Acumulación'),
                         ('J2', 'Autonómico Sub-18 · sábado 20/06')]:
        escribir(ws, celda, texto, f_dato, r_crema, None, borde)

    escribir(ws, 'A3',
             'Rellena solo las casillas de fondo crema. Una línea por ejercicio. '
             'Repite el día en todas sus líneas, y el nombre de «Para quién» en todas las suyas.   ·   '
             'Esta cabecera es para ti y para imprimir: de aquí arriba, el portal todavía no se '
             'guarda el bloque de la temporada ni la competición.',
             f_pista)

    for i, (cab, ancho) in enumerate(COLUMNAS, start=1):
        letra = get_column_letter(i)
        ws.column_dimensions[letra].width = ancho
        escribir(ws, letra + '5', cab, f_cab, r_azul, Alignment(vertical='center', wrap_text=True))
    ws.row_dimensions[5].height = 30

    for n, fila in enumerate(EJEMPLO):
        r = FILA_1 + n
        for i, valor in enumerate(fila, start=1):
            letra = get_column_letter(i)
            escribir(ws, letra + str(r), valor if valor != '' else None,
                     f_ejemplo, r_crema, arriba, borde)
    escribir(ws, get_column_letter(len(COLUMNAS) + 1) + str(FILA_1),
             '◄ EJEMPLO · bórralo y escribe encima', f_marca)

    # Filas vacías, ya con su formato: que se vea hasta dónde se puede escribir.
    for r in range(FILA_1 + len(EJEMPLO), FILA_N + 1):
        for i in range(1, len(COLUMNAS) + 1):
            escribir(ws, get_column_letter(i) + str(r), None, f_dato, r_crema, arriba, borde)

    ws.freeze_panes = 'A6'
    ws.auto_filter.ref = 'A5:%s%d' % (get_column_letter(len(COLUMNAS)), FILA_N)

    def lista(col, valores, titulo, texto):
        dv = DataValidation(type='list', formula1='"%s"' % ','.join(valores), allow_blank=True,
                            showErrorMessage=False)
        dv.promptTitle, dv.prompt, dv.showInputMessage = titulo, texto, True
        ws.add_data_validation(dv)
        dv.add('%s%d:%s%d' % (col, FILA_1, col, FILA_N))

    lista(COL['Día'], DIAS, 'Día',
          'Repite el día en cada línea de esa sesión.')
    lista(COL['Tipo de sesión'], TIPOS, 'Tipo de sesión',
          'Solo hace falta en la primera línea del día.')
    lista(COL['Papel del día'], PAPELES, 'Papel del día',
          'Qué papel juega la sesión en la semana. Solo en la primera línea del día.')
    lista(COL['Bloque'], BLOQUES, 'Bloque',
          'La parte de la sesión. Las líneas seguidas con el mismo bloque van juntas.')
    lista(COL['Calzado'], CALZADOS, 'Calzado',
          'Z = zapatillas · C = clavos · ZG = zapatillas de gimnasio.')

    def aviso(col, titulo, texto):
        dv = DataValidation(type='textLength', operator='greaterThan', formula1='0',
                            allow_blank=True, showErrorMessage=False)
        dv.promptTitle, dv.prompt, dv.showInputMessage = titulo, texto, True
        ws.add_data_validation(dv)
        dv.add('%s%d:%s%d' % (col, FILA_1, col, FILA_N))

    aviso(COL['Para quién (atleta o subgrupo)'], 'Para quién',
          'Vacío = lo hace todo el grupo.\n'
          'Con un nombre («Juan») o un subgrupo («Vallistas») = solo eso.\n'
          'Escríbelo IGUAL en todas las líneas de esa persona: si cambia una letra, '
          'salen dos bloques distintos.')
    aviso(COL['Ritmo'], 'Ritmo',
          'Si pones un porcentaje (90% o 90-95%) y la distancia en metros, cada atleta verá '
          'SU tiempo objetivo, calculado con su mejor marca.\n'
          'Si pones un ritmo fijo (7.9 s, 4:10 min/km), todos verán lo mismo.')
    aviso(COL['Distancia (m)'], 'Distancia',
          'En pista, los metros de cada repetición: 120.\n'
          'En gimnasio, las repeticiones: 3 series x 6 -> Series 3 y aquí 6.')
    aviso(COL['Series'], 'Series',
          'Solo el número. «2x120m» se parte en Series 2 y Distancia 120.')

    return ws


def hoja_repaso(wb):
    """Lo que se calcula solo. No suma metros a propósito: en atletismo el
    volumen no es la carga, y la columna de distancia son repeticiones cuando el
    día es de gimnasio. Lo que sí revienta una semana es subirla con días sin
    título o sin tipo, y eso es lo que se mira aquí."""
    ws = wb.create_sheet('Repaso')
    ws.sheet_view.showGridLines = False

    f_titulo = Font(name='Arial', bold=True, size=14, color=AZUL)
    f_pista  = Font(name='Arial', italic=True, size=9, color=GRIS)
    f_cab    = Font(name='Arial', bold=True, size=10, color='FFFFFF')
    f_dato   = Font(name='Arial', size=10)
    f_aviso  = Font(name='Arial', size=10, color=AMBAR2)
    r_azul   = PatternFill('solid', fgColor=AZUL)
    r_cielo  = PatternFill('solid', fgColor=CIELO)
    r_crema  = PatternFill('solid', fgColor=CREMA)

    for i, ancho in enumerate([16, 14, 46, 22], start=1):
        ws.column_dimensions[get_column_letter(i)].width = ancho

    escribir(ws, 'A1', 'LO QUE HAY DE VERDAD', f_titulo)
    escribir(ws, 'A2', 'Se calcula solo desde la hoja «Semana». Míralo antes de subir el archivo.', f_pista)

    dia, tit, tip, ejer, quien = (COL['Día'], COL['Título de la sesión'], COL['Tipo de sesión'],
                                  COL['Ejercicio'], COL['Para quién (atleta o subgrupo)'])
    rango = lambda c: "Semana!$%s$%d:$%s$%d" % (c, FILA_1, c, FILA_N)

    for i, cab in enumerate(['Día', 'Ejercicios', 'Le falta'], start=1):
        escribir(ws, get_column_letter(i) + '4', cab, f_cab, r_azul)
    for n, d in enumerate(DIAS):
        r = 5 + n
        escribir(ws, 'A%d' % r, d, f_dato, r_cielo)
        escribir(ws, 'B%d' % r, '=COUNTIFS(%s,$A%d,%s,"<>")' % (rango(dia), r, rango(ejer)), f_dato, r_cielo)
        # Un día con ejercicios pero sin título ni tipo entra como sesión sin nombre.
        escribir(ws, 'C%d' % r,
                 '=IF(B{r}=0,"",'
                 'IF(AND(COUNTIFS({rd},$A{r},{rt},"<>")=0,COUNTIFS({rd},$A{r},{rp},"<>")=0),'
                 '"le falta el título y el tipo de sesión",'
                 'IF(COUNTIFS({rd},$A{r},{rt},"<>")=0,"le falta el título de la sesión",'
                 'IF(COUNTIFS({rd},$A{r},{rp},"<>")=0,"le falta el tipo de sesión",""))))'
                 .format(r=r, rd=rango(dia), rt=rango(tit), rp=rango(tip)),
                 f_aviso, r_cielo)

    escribir(ws, 'A14', 'CUÁNTO LLEVA CADA UNO', f_titulo)
    escribir(ws, 'A15',
             'Escribe abajo los nombres tal como los has puesto en «Para quién» y verás cuántos '
             'ejercicios propios lleva cada uno esta semana. Lo que hace todo el grupo no se cuenta aquí.',
             f_pista)
    for i, cab in enumerate(['Atleta o subgrupo', 'Ejercicios suyos'], start=1):
        escribir(ws, get_column_letter(i) + '17', cab, f_cab, r_azul)
    for n in range(8):
        r = 18 + n
        escribir(ws, 'A%d' % r, None, f_dato, r_crema)
        escribir(ws, 'B%d' % r,
                 '=IF($A{r}="","",COUNTIFS({rq},$A{r},{re},"<>"))'.format(r=r, rq=rango(quien), re=rango(ejer)),
                 f_dato, r_cielo)
    return ws


def hoja_ayuda(wb):
    ws = wb.create_sheet('Cómo se rellena')
    ws.sheet_view.showGridLines = False
    ws.column_dimensions['A'].width = 112

    f_h1 = Font(name='Arial', bold=True, size=14, color=AZUL)
    f_h2 = Font(name='Arial', bold=True, size=11, color=AZUL)
    f_p  = Font(name='Arial', size=10)
    f_av = Font(name='Arial', bold=True, size=10, color=AMBAR2)
    izq  = Alignment(vertical='top', wrap_text=True)

    lineas = [
        ('h1', 'CÓMO SE RELLENA'),
        ('p',  'Una línea por ejercicio. El día se repite en todas sus líneas. El título, la nota, la hora, '
               'el lugar, el tipo y el papel del día solo hacen falta en la PRIMERA línea de cada día.'),
        ('', ''),
        ('h2', 'Cómo se parte lo que ya escribes'),
        ('p',  'Lo que hoy escribes de corrido en una celda se reparte en columnas. Es el mismo dato:'),
        ('p',  '    2x120m @92–94% // 6\'      →   Series 2 · Distancia 120 · Ritmo 92–94% · Rec 6\''),
        ('p',  '    4x20m desde tacos @95% // 3\'   →   Ejercicio «Salidas desde tacos» · Series 4 · '
               'Distancia 20 · Ritmo 95% · Rec 3\''),
        ('p',  '    3x6 al 75% RM   →   Series 3 · Distancia 6 (las repeticiones) · Carga 75% RM'),
        ('', ''),
        ('h2', 'Para quién'),
        ('p',  'Es la columna que hace que una misma semana valga para gente que entrena junta pero '
               'no hace lo mismo.'),
        ('p',  '    Vacío  = lo hace todo el grupo.'),
        ('p',  '    Un nombre («Juan») = solo esa persona.'),
        ('p',  '    Un subgrupo («Vallistas») = ese grupo de dentro del grupo.'),
        ('av', 'Escríbelo IGUAL en todas las líneas de esa persona. «Juan» y «juan » son dos bloques '
               'distintos, y el portal los separa como los escribas.'),
        ('', ''),
        ('h2', 'Ritmo: por qué merece la pena ponerlo en %'),
        ('p',  'Si pones un porcentaje (90%, 90-95%) y la distancia en metros, cada atleta ve SU tiempo '
               'objetivo, calculado con su mejor marca en esa distancia. Si pones un ritmo fijo (7.9 s, '
               '4:10 min/km), todos ven exactamente lo mismo.'),
        ('', ''),
        ('h2', 'Calzado'),
        ('p',  'Z = zapatillas · C = clavos · ZG = zapatillas de gimnasio · Descalzo.'),
        ('', ''),
        ('h2', 'La cabecera de arriba'),
        ('p',  'El grupo, la semana, el bloque de la temporada y la competición de referencia se '
               'escriben porque una planificación sin eso no se entiende dentro de un mes. Pero de '
               'momento se quedan en la hoja: el portal guarda los días y los ejercicios, no la '
               'cabecera de la semana. Si necesitas que algo de ahí lo vea el atleta, ponlo en la '
               'nota del día.'),
        ('', ''),
        ('h2', 'Días que no se entrena'),
        ('p',  'Se ponen también: una línea con el día, el título «Descanso» y el tipo «Descanso». '
               'Nada más.'),
        ('', ''),
        ('h1', 'CÓMO SE GUARDA Y DÓNDE SE SUBE'),
        ('p',  '1. En Drive: se sube este archivo y se abre con Hojas de cálculo. Las listas y las '
               'fórmulas siguen funcionando.'),
        ('p',  '2. Se rellena la hoja «Semana» y se mira la hoja «Repaso».'),
        ('p',  '3. Archivo → Descargar → Valores separados por comas (.csv).'),
        ('av', '⚠️ Drive descarga SOLO LA HOJA QUE TENGAS ABIERTA. Ponte en «Semana» antes de '
               'descargar, o bajarás la de ayuda y el portal no entenderá nada.'),
        ('p',  '4. En el portal: Planificar semana → Montar la semana de una vez → 2 · Importar un '
               'archivo. Sale una previsualización y NO SE GUARDA NADA hasta que le das a crear.'),
        ('', ''),
        ('h2', 'Si prefieres pedírselo a una IA'),
        ('p',  'Mándale esta misma hoja y dile que te devuelva la semana con estas columnas, en el '
               'mismo orden. Luego la pegas en la hoja «Semana» y la corriges tú, que es lo que de '
               'verdad lleva tiempo.'),
        ('', ''),
        ('h2', 'Si cambias las cabeceras'),
        ('av', 'No las cambies. El portal reconoce cada columna por su nombre: si le cambias el título '
               'a una, esa columna entera se pierde al subir el archivo y no avisa nadie.'),
    ]
    r = 1
    for tipo, texto in lineas:
        if tipo:
            c = escribir(ws, 'A%d' % r, texto,
                         {'h1': f_h1, 'h2': f_h2, 'p': f_p, 'av': f_av}[tipo], None, izq)
            if tipo == 'h1':
                ws.row_dimensions[r].height = 24
        r += 1
    return ws


def main():
    wb = Workbook()
    wb.remove(wb.active)
    hoja_semana(wb)
    hoja_repaso(wb)
    hoja_ayuda(wb)
    # Sin LibreOffice no se pueden dejar las fórmulas calculadas, así que se le
    # pide a Excel y a Sheets que las calculen al abrir. Solo se usan COUNTIFS,
    # IF y AND, que las entienden los tres.
    wb.calculation.fullCalcOnLoad = True

    raiz = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    destino = os.path.join(raiz, 'assets', 'descargas', 'plantilla-semana-atletismo.xlsx')
    wb.save(destino)
    print('Escrito: ' + destino)


if __name__ == '__main__':
    main()
